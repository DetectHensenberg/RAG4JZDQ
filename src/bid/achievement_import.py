"""Excel / CSV batch import for achievement records.

Parses uploaded files and maps columns to :class:`AchievementRecord` fields.
Supports ``.xlsx`` (via openpyxl) and ``.csv`` (stdlib).
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.bid.achievement_db import AchievementRecord, create_achievement

logger = logging.getLogger(__name__)

# Column aliases → canonical field name
_COLUMN_MAP: Dict[str, str] = {
    "项目名称": "project_name",
    "project_name": "project_name",
    "项目内容": "project_content",
    "project_content": "project_content",
    "项目金额": "amount",
    "合同金额": "amount",
    "金额": "amount",
    "amount": "amount",
    "合同签订时间": "sign_date",
    "签订时间": "sign_date",
    "签约日期": "sign_date",
    "sign_date": "sign_date",
    "验收时间": "acceptance_date",
    "验收日期": "acceptance_date",
    "acceptance_date": "acceptance_date",
    "甲方联系人": "client_contact",
    "联系人": "client_contact",
    "client_contact": "client_contact",
    "联系方式": "client_phone",
    "电话": "client_phone",
    "client_phone": "client_phone",
    "标签": "tags",
    "tags": "tags",
}


@dataclass
class ImportResult:
    """Result of a batch import operation."""

    success_count: int = 0
    error_count: int = 0
    errors: List[str] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


def _normalize_header(h: str) -> Optional[str]:
    """Map a raw header string to a canonical field name."""
    return _COLUMN_MAP.get(h.strip())


def _parse_amount(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        s = str(val).replace(",", "").replace("，", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_tags(val: Any) -> List[str]:
    if not val:
        return []
    s = str(val).strip()
    for sep in [",", "，", ";", "；", "|", "、"]:
        if sep in s:
            return [t.strip() for t in s.split(sep) if t.strip()]
    return [s] if s else []


def _row_to_record(row: Dict[str, Any], field_map: Dict[int, str], idx: int) -> tuple[Optional[AchievementRecord], Optional[str]]:
    """Convert a mapped row dict to an AchievementRecord.

    Returns ``(record, None)`` on success or ``(None, error_msg)`` on failure.
    """
    project_name = str(row.get("project_name", "")).strip()
    if not project_name:
        return None, f"第 {idx} 行: 项目名称为空"

    return AchievementRecord(
        project_name=project_name,
        project_content=str(row.get("project_content", "")).strip(),
        amount=_parse_amount(row.get("amount")),
        sign_date=str(row.get("sign_date", "")).strip(),
        acceptance_date=str(row.get("acceptance_date", "")).strip(),
        client_contact=str(row.get("client_contact", "")).strip(),
        client_phone=str(row.get("client_phone", "")).strip(),
        tags=_parse_tags(row.get("tags")),
    ), None


def import_csv(content: bytes, encoding: str = "utf-8-sig") -> ImportResult:
    """Import achievements from CSV bytes."""
    result = ImportResult()
    try:
        text = content.decode(encoding)
    except UnicodeDecodeError:
        text = content.decode("gbk", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        result.errors.append("CSV 无表头")
        return result

    field_map = {}
    for i, h in enumerate(reader.fieldnames):
        canonical = _normalize_header(h)
        if canonical:
            field_map[i] = canonical

    if "project_name" not in field_map.values():
        result.errors.append("CSV 缺少「项目名称」列")
        return result

    for row_num, row in enumerate(reader, start=2):
        mapped = {}
        for raw_key, value in row.items():
            canonical = _normalize_header(raw_key)
            if canonical:
                mapped[canonical] = value

        record, err = _row_to_record(mapped, field_map, row_num)
        if err:
            result.errors.append(err)
            result.error_count += 1
            continue

        try:
            create_achievement(record)  # type: ignore[arg-type]
            result.success_count += 1
        except Exception as e:
            result.errors.append(f"第 {row_num} 行写入失败: {e}")
            result.error_count += 1

    return result


def import_excel(content: bytes) -> ImportResult:
    """Import achievements from an Excel (.xlsx) file."""
    result = ImportResult()
    try:
        from openpyxl import load_workbook
    except ImportError:
        result.errors.append("openpyxl 未安装，无法解析 Excel 文件")
        return result

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        result.errors.append("Excel 文件无活动工作表")
        return result

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        result.errors.append("Excel 文件为空")
        return result

    field_map: Dict[int, str] = {}
    for i, cell_val in enumerate(header_row):
        if cell_val is not None:
            canonical = _normalize_header(str(cell_val))
            if canonical:
                field_map[i] = canonical

    if "project_name" not in field_map.values():
        result.errors.append("Excel 缺少「项目名称」列")
        return result

    for row_num, row_vals in enumerate(rows_iter, start=2):
        mapped: Dict[str, Any] = {}
        for col_idx, canonical in field_map.items():
            if col_idx < len(row_vals):
                mapped[canonical] = row_vals[col_idx]

        record, err = _row_to_record(mapped, field_map, row_num)
        if err:
            result.errors.append(err)
            result.error_count += 1
            continue

        try:
            create_achievement(record)  # type: ignore[arg-type]
            result.success_count += 1
        except Exception as e:
            result.errors.append(f"第 {row_num} 行写入失败: {e}")
            result.error_count += 1

    wb.close()
    return result


def import_file(filename: str, content: bytes) -> ImportResult:
    """Auto-detect file type and import."""
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return import_csv(content)
    elif ext in (".xlsx", ".xls"):
        return import_excel(content)
    else:
        r = ImportResult()
        r.errors.append(f"不支持的文件类型: {ext}，仅支持 .csv / .xlsx")
        return r
